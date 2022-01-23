#virtualenv -p ~/anaconda3/envs/solina_bot_32/python.exe

import time
import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd
from openpyxl import load_workbook
import shutil
import xlwings as xw
print("aggregation")
# reading stocks code list
file = open("tradebot_codes.txt", "r")
codes = file.read().splitlines()

today = datetime.now()
today = today.strftime("%Y%m%d")
source_file = 'C:/OpenAPI/kiwoom_tradebot/solina_bot_data/' + today + '/data.xlsx'
destination = "C:/OpenAPI/kiwoom_tradebot/solina_bot_data/final_sheets/"
origin = "C:/OpenAPI/kiwoom_tradebot/reference_file"

# reading column headers from reference file
df = pd.read_excel(origin + '/reference_file.xlsx', sheet_name = 'RAWDATA', engine = 'openpyxl')
df = df[0:0]

# 여기서부터 종목 코드 별 진행
for c in codes:
    # reference file 복사
    file_name = today + '_' + c + '_final.xlsx'
    file_path = destination + file_name
    
    files = os.listdir(origin)
    if not os.path.isdir(destination):
        os.makedirs(destination)
    for file in files:
        if not os.path.exists(file_path):
            shutil.copy(origin + '/reference_file.xlsx', file_path)
            print("success: reference file copied")
    
    wb = xw.Book(file_path)
    ws = wb.sheets["RAWDATA"]
    

    sheet_name = '10059_' + c
    df1 = pd.read_excel(source_file, sheet_name=sheet_name, engine = 'openpyxl')
    sheet_name = '10047_' + c
    df2 = pd.read_excel(source_file, sheet_name=sheet_name, engine = 'openpyxl')

    del df1["대비기호"]
    del df1["등락률"]
    del df1["누적거래대금"]

    df1.insert(3, "체결강도", df2["체결강도"], True)
    df1.insert(8, "", 0, True)
    #print(df1)
    #df.append(df1, ignore_index = True)
    #print(df)
    ws.range('A1').value = df1
    ws.range('A1').expand().value
#    ws["A1"].options(pd.DataFrame, header=1, index=True, expand='table').values = df1
    wb.save(file_path)
    wb.close()
    #df1.to_excel(file_path, index = False)

os.system("taskkill /f /im Excel.exe")
print("success. complete.")



